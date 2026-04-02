"""
Generate auditable Excel templates with named ranges for standard RE deliverables.

Run this script to create/update the template files in assets/.
Templates have standard structure, named ranges, and formatting that the LLM fills with data.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ASSETS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets")

HEADER_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="444444")
NUMBER_FONT = Font(name="Calibri", size=10)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="999999")
)
DOLLAR_FMT = '#,##0'
DOLLAR_CENTS_FMT = '#,##0.00'
PCT_FMT = '0.00%'
PSF_FMT = '#,##0.00'


def create_proforma_template():
    """Create the standard pro forma Excel template."""
    wb = Workbook()

    # --- Tab 1: Assumptions ---
    ws = wb.active
    ws.title = "Assumptions"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    ws['A1'] = "PROPERTY UNDERWRITING ASSUMPTIONS"
    ws['A1'].font = TITLE_FONT

    sections = [
        ("Property Information", [
            ("Property Name", "property_name", ""),
            ("Address", "property_address", ""),
            ("Property Type", "property_type", ""),
            ("Total SF / Units", "total_size", 0),
            ("Year Built", "year_built", 0),
            ("Purchase Price", "purchase_price", 0),
            ("Closing Costs %", "closing_costs_pct", 0.02),
        ]),
        ("Revenue Assumptions", [
            ("In-Place Occupancy", "occupancy", 0),
            ("Market Rent PSF", "market_rent_psf", 0),
            ("Annual Rent Growth", "rent_growth", 0.03),
            ("Structural Vacancy", "vacancy_rate", 0.05),
            ("Credit Loss", "credit_loss_pct", 0.01),
        ]),
        ("Expense Assumptions", [
            ("Year 1 Operating Expenses", "base_expenses", 0),
            ("Expense Growth Rate", "expense_growth", 0.03),
            ("Management Fee %", "mgmt_fee_pct", 0.03),
        ]),
        ("Capital Assumptions", [
            ("Capital Reserves PSF", "capex_reserves_psf", 0.25),
            ("TI PSF (New)", "ti_new_psf", 0),
            ("TI PSF (Renewal)", "ti_renewal_psf", 0),
            ("LC % (New)", "lc_new_pct", 0),
            ("LC % (Renewal)", "lc_renewal_pct", 0),
        ]),
        ("Debt Assumptions", [
            ("Loan Amount", "loan_amount", 0),
            ("Interest Rate", "interest_rate", 0),
            ("Amortization (Years)", "amort_years", 30),
            ("IO Period (Years)", "io_years", 0),
            ("Loan Term (Years)", "loan_term", 10),
        ]),
        ("Exit Assumptions", [
            ("Hold Period (Years)", "hold_period", 5),
            ("Exit Cap Rate", "exit_cap_rate", 0.06),
            ("Selling Costs %", "selling_costs_pct", 0.02),
        ]),
    ]

    row = 3
    for section_name, items in sections:
        ws.cell(row=row, column=1, value=section_name).font = SUBTITLE_FONT
        row += 1
        for label, name, default in items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=default)
            # Define named range
            wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!$B${row}"))
            if isinstance(default, float) and default < 1:
                cell.number_format = PCT_FMT
            elif isinstance(default, (int, float)) and default >= 1:
                cell.number_format = DOLLAR_FMT
            row += 1
        row += 1

    # --- Tab 2: Cash Flow Summary ---
    ws2 = wb.create_sheet("Cash Flow Summary")
    ws2.column_dimensions['A'].width = 30

    ws2['A1'] = "CASH FLOW SUMMARY"
    ws2['A1'].font = TITLE_FONT

    # Header row with years
    year_labels = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5",
                   "Year 6", "Year 7", "Year 8", "Year 9", "Year 10"]
    for col, label in enumerate(year_labels, 1):
        cell = ws2.cell(row=3, column=col, value=label)
        if col > 1:
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col)].width = 14

    # Line items
    line_items = [
        ("REVENUE", None, True),
        ("Potential Gross Income", "pgi", False),
        ("Less: Vacancy", "vacancy", False),
        ("Less: Credit Loss", "credit_loss", False),
        ("Effective Gross Income", "egi", True),
        ("", None, False),
        ("EXPENSES", None, True),
        ("Operating Expenses", "opex", False),
        ("Management Fee", "mgmt_fee", False),
        ("Total Expenses", "total_expenses", True),
        ("", None, False),
        ("NET OPERATING INCOME", "noi", True),
        ("", None, False),
        ("BELOW THE LINE", None, True),
        ("Capital Reserves", "capex_reserves", False),
        ("Tenant Improvements", "ti_cost", False),
        ("Leasing Commissions", "lc_cost", False),
        ("Net Cash Flow", "ncf", True),
        ("", None, False),
        ("Debt Service", "debt_service", False),
        ("Cash Flow After Debt", "cfad", True),
    ]

    row = 4
    for label, name, is_subtotal in line_items:
        cell = ws2.cell(row=row, column=1, value=label)
        if is_subtotal:
            cell.font = HEADER_FONT
            for col in range(1, 12):
                ws2.cell(row=row, column=col).fill = SUBTOTAL_FILL
                ws2.cell(row=row, column=col).border = THIN_BORDER
        if name:
            for yr in range(1, 11):
                c = ws2.cell(row=row, column=yr + 1, value=0)
                c.number_format = DOLLAR_FMT
                c.alignment = Alignment(horizontal="right")
            # Named range for Year 1 of each line item
            wb.defined_names.add(DefinedName(f"cf_{name}_y1", attr_text=f"'Cash Flow Summary'!$B${row}"))
        row += 1

    # --- Tab 3: Returns Summary ---
    ws3 = wb.create_sheet("Returns Summary")
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20

    ws3['A1'] = "RETURNS SUMMARY"
    ws3['A1'].font = TITLE_FONT

    ws3.cell(row=3, column=1, value="Metric").font = HEADER_FONT_WHITE
    ws3.cell(row=3, column=1).fill = HEADER_FILL
    ws3.cell(row=3, column=2, value="Unlevered").font = HEADER_FONT_WHITE
    ws3.cell(row=3, column=2).fill = HEADER_FILL
    ws3.cell(row=3, column=3, value="Levered").font = HEADER_FONT_WHITE
    ws3.cell(row=3, column=3).fill = HEADER_FILL

    metrics = [
        ("Going-In Cap Rate", "going_in_cap", PCT_FMT),
        ("Exit Cap Rate", "exit_cap", PCT_FMT),
        ("IRR", "irr", PCT_FMT),
        ("Equity Multiple", "equity_multiple", "0.00x"),
        ("Cash-on-Cash (Yr 1)", "coc_y1", PCT_FMT),
        ("Average Annual Cash Yield", "avg_cash_yield", PCT_FMT),
        ("Price / SF", "price_psf", PSF_FMT),
        ("Price / Unit", "price_per_unit", DOLLAR_FMT),
    ]

    row = 4
    for label, name, fmt in metrics:
        ws3.cell(row=row, column=1, value=label)
        for col in [2, 3]:
            c = ws3.cell(row=row, column=col, value=0)
            c.number_format = fmt
        wb.defined_names.add(DefinedName(f"ret_{name}_unlev", attr_text=f"'Returns Summary'!$B${row}"))
        wb.defined_names.add(DefinedName(f"ret_{name}_lev", attr_text=f"'Returns Summary'!$C${row}"))
        row += 1

    # --- Tab 4: Sources & Assumptions ---
    ws4 = wb.create_sheet("Sources & Assumptions")
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 20

    ws4['A1'] = "SOURCES & ASSUMPTIONS"
    ws4['A1'].font = TITLE_FONT

    ws4.cell(row=3, column=1, value="Data Item").font = HEADER_FONT_WHITE
    ws4.cell(row=3, column=1).fill = HEADER_FILL
    ws4.cell(row=3, column=2, value="Source").font = HEADER_FONT_WHITE
    ws4.cell(row=3, column=2).fill = HEADER_FILL
    ws4.cell(row=3, column=3, value="As-Of Date").font = HEADER_FONT_WHITE
    ws4.cell(row=3, column=3).fill = HEADER_FILL

    source_items = [
        "Rent Roll", "Trailing Financials (T-12)", "Market Rents",
        "Cap Rate", "Expense Growth", "Vacancy Assumption",
        "Comparable Sales", "Debt Terms", "Exit Cap Rate",
    ]
    for i, item in enumerate(source_items, 4):
        ws4.cell(row=i, column=1, value=item)
        ws4.cell(row=i, column=2, value="[To be completed]")
        ws4.cell(row=i, column=3, value="[Date]")

    # Save
    os.makedirs(ASSETS_DIR, exist_ok=True)
    filepath = os.path.join(ASSETS_DIR, "proforma-template.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


def create_debt_sizing_template():
    """Create the debt sizing comparison template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Debt Sizing"
    ws.column_dimensions['A'].width = 28

    ws['A1'] = "DEBT SIZING ANALYSIS"
    ws['A1'].font = TITLE_FONT

    # Lender comparison columns
    headers = ["", "Scenario 1", "Scenario 2", "Scenario 3"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        if col > 1:
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(col)].width = 18

    items = [
        ("PROPERTY", None),
        ("NOI", DOLLAR_FMT),
        ("Property Value", DOLLAR_FMT),
        ("", None),
        ("LOAN TERMS", None),
        ("Lender Type", None),
        ("Interest Rate", PCT_FMT),
        ("Amortization (Yrs)", "0"),
        ("IO Period (Yrs)", "0"),
        ("Term (Yrs)", "0"),
        ("", None),
        ("CONSTRAINTS", None),
        ("Max LTV", PCT_FMT),
        ("Min DSCR", "0.00x"),
        ("Min Debt Yield", PCT_FMT),
        ("", None),
        ("SIZING RESULTS", None),
        ("Max Loan (LTV)", DOLLAR_FMT),
        ("Max Loan (DSCR)", DOLLAR_FMT),
        ("Max Loan (DY)", DOLLAR_FMT),
        ("Binding Constraint", None),
        ("SIZED LOAN", DOLLAR_FMT),
        ("", None),
        ("ACTUAL METRICS", None),
        ("LTV", PCT_FMT),
        ("DSCR", "0.00x"),
        ("Debt Yield", PCT_FMT),
        ("Annual Debt Service", DOLLAR_FMT),
        ("Breakeven Occupancy", PCT_FMT),
    ]

    row = 4
    for label, fmt in items:
        cell = ws.cell(row=row, column=1, value=label)
        if fmt is None and label and label.isupper():
            cell.font = HEADER_FONT
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
        elif fmt:
            for c in range(2, 5):
                ws.cell(row=row, column=c).number_format = fmt
        row += 1

    # Sources tab
    ws2 = wb.create_sheet("Sources & Assumptions")
    ws2['A1'] = "SOURCES & ASSUMPTIONS"
    ws2['A1'].font = TITLE_FONT

    os.makedirs(ASSETS_DIR, exist_ok=True)
    filepath = os.path.join(ASSETS_DIR, "debt-sizing-template.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


if __name__ == "__main__":
    create_proforma_template()
    create_debt_sizing_template()
    print("All templates created successfully.")
